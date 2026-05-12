#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 标注器 - 动态密度扫描（降低阈值，允许间隙）
"""

import os
import json
import re
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz

logger = logging.getLogger(__name__)

def normalize_text(text):
    if text is None:
        return ""
    return re.sub(r'\s+', '', str(text))

class ExcelAnnotator:
    def __init__(self, threshold=80, method='ratio', strict=True,
                 density_min_count=1, density_min_ratio=0.1,
                 max_gap=2):
        """
        :param threshold: 模糊匹配阈值
        :param method: 匹配方法
        :param strict: 是否启用严格模式
        :param density_min_count: 表头行/列的最小匹配数（绝对数量）
        :param density_min_ratio: 表头行/列的最小匹配数占该行非空单元格的比例
        :param max_gap: 允许连续的非密集行数（用于跳过空白/非表头行）
        """
        self.threshold = threshold
        self.method = method
        self.strict = strict
        self.density_min_count = density_min_count
        self.density_min_ratio = density_min_ratio
        self.max_gap = max_gap

    def _get_data_start_row_col(self, sheet):
        min_row = None
        min_col = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip():
                    if min_row is None or cell.row < min_row:
                        min_row = cell.row
                    if min_col is None or cell.column < min_col:
                        min_col = cell.column
        if min_row is None:
            min_row = 1
        if min_col is None:
            min_col = 1
        return min_row, min_col

    def _is_numeric_header(self, text):
        clean = normalize_text(text)
        return clean.isdigit()

    def find_cell_by_text(self, sheet, text, region=None):
        matched_cells = []
        text_norm = normalize_text(text)
        if not text_norm:
            return matched_cells
        if region:
            min_row, max_row, min_col, max_col = region
            rows = range(min_row, max_row + 1)
            cols = range(min_col, max_col + 1)
        else:
            rows = range(1, sheet.max_row + 1)
            cols = range(1, sheet.max_column + 1)
        for row in rows:
            for col in cols:
                cell = sheet.cell(row, col)
                if not cell.value:
                    continue
                cell_str = normalize_text(cell.value)
                if not cell_str:
                    continue
                if self.method == 'ratio':
                    sim = fuzz.ratio(cell_str, text_norm)
                elif self.method == 'partial_ratio':
                    sim = fuzz.partial_ratio(cell_str, text_norm)
                elif self.method == 'token_sort_ratio':
                    sim = fuzz.token_sort_ratio(cell_str, text_norm)
                else:
                    sim = fuzz.ratio(cell_str, text_norm)
                if sim >= self.threshold:
                    matched_cells.append(cell)
        return matched_cells

    def annotate_excel(self, excel_path, json_path, output_path):
        with open(json_path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
        annotations = data.get('headers', [])
        logger.info(f"处理文件: {excel_path}")
        logger.info(f"JSON中包含 {len(annotations)} 条标注")

        wb = load_workbook(excel_path)
        sheet = wb.active
        max_row = sheet.max_row
        max_col = sheet.max_column

        # 收集候选
        col_candidates = []
        row_candidates = []
        for anno in annotations:
            words = anno.get('words', '')
            cell_type = anno.get('type', '')
            if not words or not cell_type:
                continue
            if cell_type not in ('column_header', 'row_header'):
                continue
            matched_cells = self.find_cell_by_text(sheet, words, region=None)
            if not matched_cells:
                logger.warning(f"未找到匹配单元格: [{cell_type}] '{words}'")
                continue
            if cell_type == 'column_header':
                col_candidates.append((words, cell_type, matched_cells))
            else:
                row_candidates.append((words, cell_type, matched_cells))

        # 动态确定列头行区域（允许间隙）
        col_header_rows = set()
        if col_candidates:
            row_match_count = {}
            for _, _, cells in col_candidates:
                matched_rows = set(cell.row for cell in cells)
                for row in matched_rows:
                    row_match_count[row] = row_match_count.get(row, 0) + 1
            max_scan_rows = min(max_row, 200)
            start_row = None
            end_row = None
            gap = 0
            for row in range(1, max_scan_rows + 1):
                count = row_match_count.get(row, 0)
                non_empty = sum(1 for col in range(1, max_col+1) if sheet.cell(row, col).value)
                ratio = count / non_empty if non_empty > 0 else 0
                is_dense = (count >= self.density_min_count) or (ratio >= self.density_min_ratio)
                if is_dense:
                    if start_row is None:
                        start_row = row
                    end_row = row
                    gap = 0
                else:
                    if start_row is not None:
                        gap += 1
                        if gap > self.max_gap:
                            break
            if start_row is not None:
                col_header_rows = set(range(start_row, end_row + 1))
                logger.info(f"动态识别列头区域: 行 {col_header_rows}")
            else:
                logger.warning("未检测到列头区域，将使用所有匹配")

        # 动态确定行头列区域
        row_header_cols = set()
        if row_candidates:
            col_match_count = {}
            for _, _, cells in row_candidates:
                matched_cols = set(cell.column for cell in cells)
                for col in matched_cols:
                    col_match_count[col] = col_match_count.get(col, 0) + 1
            max_scan_cols = min(max_col, 100)
            start_col = None
            end_col = None
            gap = 0
            for col in range(1, max_scan_cols + 1):
                count = col_match_count.get(col, 0)
                non_empty = sum(1 for row in range(1, max_row+1) if sheet.cell(row, col).value)
                ratio = count / non_empty if non_empty > 0 else 0
                is_dense = (count >= self.density_min_count) or (ratio >= self.density_min_ratio)
                if is_dense:
                    if start_col is None:
                        start_col = col
                    end_col = col
                    gap = 0
                else:
                    if start_col is not None:
                        gap += 1
                        if gap > self.max_gap:
                            break
            if start_col is not None:
                row_header_cols = set(range(start_col, end_col + 1))
                logger.info(f"动态识别行头区域: 列 {row_header_cols}")
            else:
                logger.warning("未检测到行头区域，将使用所有匹配")

        # 标色
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        matched_stats = {'column_header': 0, 'row_header': 0, 'total_cells_colored': 0}

        for words, cell_type, cells in col_candidates:
            fill = red_fill
            for cell in cells:
                if col_header_rows and cell.row not in col_header_rows:
                    continue
                cell.fill = fill
                matched_stats['column_header'] += 1
                matched_stats['total_cells_colored'] += 1
                logger.info(f"标色列头: '{words}' -> ({cell.row},{cell.column})")

        for words, cell_type, cells in row_candidates:
            fill = green_fill
            for cell in cells:
                if row_header_cols and cell.column not in row_header_cols:
                    continue
                cell.fill = fill
                matched_stats['row_header'] += 1
                matched_stats['total_cells_colored'] += 1
                logger.info(f"标色行头: '{words}' -> ({cell.row},{cell.column})")

        wb.save(output_path)
        logger.info(f"已保存标注文件: {output_path}")
        logger.info(f"统计: 列头 {matched_stats['column_header']}, 行头 {matched_stats['row_header']}, 总标色 {matched_stats['total_cells_colored']}")

    def batch_annotate(self, table_folder, vlm_folder, output_folder):
        os.makedirs(output_folder, exist_ok=True)
        excel_files = [f for f in os.listdir(table_folder) if f.endswith('.xlsx')]
        if not excel_files:
            print("❌ 未找到 Excel 文件")
            return
        for filename in excel_files:
            base = os.path.splitext(filename)[0]
            excel_path = os.path.join(table_folder, filename)
            vlm_json_path = os.path.join(vlm_folder, f"{base}_e2e_vlm.json")
            if not os.path.exists(vlm_json_path):
                print(f"⚠️ 跳过 {filename}：未找到对应的 VLM 结果文件 {vlm_json_path}")
                continue
            output_path = os.path.join(output_folder, filename)
            print(f"\n🔧 处理 {filename} ...")
            self.annotate_excel(excel_path, vlm_json_path, output_path)