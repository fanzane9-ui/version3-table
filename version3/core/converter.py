#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
直接表头转图片转换器 - VBA宏增强版，支持格式标准化（居中、边框增强、字体增强）及自动跳过顶部嵌入图片
"""

import os
import time
import openpyxl
from PIL import Image, ImageGrab
import xlwings as xw
import win32clipboard
import struct
from io import BytesIO

XLWINGS_AVAILABLE = True


class DirectHeaderConverter:
    def __init__(self, apply_formatting=True, verbose=False):
        self.app = None
        self.apply_formatting = apply_formatting
        self.verbose = verbose
        self.vba_macro_code = '''
Option Explicit

Public Sub AutoFitAllMergedCells()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets(1)
    If ws Is Nothing Then Exit Sub

    Dim cell As Range
    Dim processed As New Collection
    Dim mergedArea As Range
    Dim addr As String

    On Error Resume Next
    For Each cell In ws.UsedRange
        If cell.MergeCells Then
            Set mergedArea = cell.MergeArea
            addr = mergedArea.Address
            processed.Add addr, addr
            If Err.Number = 0 Then
                Call AdjustMergedCellHeight(mergedArea)
            End If
            Err.Clear
        End If
    Next cell
    On Error GoTo 0
End Sub

Private Sub AdjustMergedCellHeight(ByVal oRange As Range)
    Dim i As Integer
    Dim totalColWidth As Single
    Dim oldZZWidth As Single
    Dim needHeight As Single
    Dim currentHeight As Single
    Dim r As Integer

    With oRange.Worksheet
        totalColWidth = 0
        For i = 1 To oRange.Columns.Count
            totalColWidth = totalColWidth + .Columns(oRange.Column + i - 1).ColumnWidth
        Next i

        Dim content As String
        content = oRange.Cells(1, 1).Value
        If Len(Trim(content)) = 0 Then Exit Sub

        oldZZWidth = .Columns("ZZ").ColumnWidth
        .Range("ZZ1").Value = content
        .Range("ZZ1").WrapText = True
        .Columns("ZZ").ColumnWidth = totalColWidth
        .Rows(1).AutoFit
        needHeight = .Rows(1).RowHeight

        currentHeight = 0
        For r = oRange.Row To oRange.Row + oRange.Rows.Count - 1
            currentHeight = currentHeight + .Rows(r).RowHeight
        Next r

        If needHeight > currentHeight Then
            Dim eachRowHeight As Single
            eachRowHeight = needHeight / oRange.Rows.Count
            For r = oRange.Row To oRange.Row + oRange.Rows.Count - 1
                .Rows(r).RowHeight = eachRowHeight
            Next r
        End If

        .Range("ZZ1").ClearContents
        .Columns("ZZ").ColumnWidth = oldZZWidth
    End With
End Sub
'''

    def __enter__(self):
        try:
            self.app = xw.App(visible=True, add_book=False)
            self.app.display_alerts = False
            self.app.screen_updating = True
        except Exception as e:
            print(f"Excel 启动失败: {e}")
            self.app = None
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.app:
            try:
                self.app.quit()
            except:
                pass

    def _get_clipboard_image(self):
        try:
            img = ImageGrab.grabclipboard()
            if img and isinstance(img, Image.Image) and img.size[0] > 10 and img.size[1] > 10:
                return img
        except:
            pass
        try:
            win32clipboard.OpenClipboard()
            if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_DIB):
                data = win32clipboard.GetClipboardData(win32clipboard.CF_DIB)
                header = struct.pack('<2sIHHII', b'BM', len(data) + 14, 0, 0, 14 + 40, 0)
                bmp_data = header + data
                img = Image.open(BytesIO(bmp_data))
                return img
        except Exception as e:
            if self.verbose:
                print(f"win32 获取失败: {e}")
        finally:
            try:
                win32clipboard.CloseClipboard()
            except:
                pass
        return None

    def _col_num_to_letter(self, n):
        result = ""
        while n > 0:
            n -= 1
            result = chr(n % 26 + ord('A')) + result
            n //= 26
        return result

    def _get_min_row_skip_images(self, ws_com, data_min_row):
        """
        检测工作表中的图片（Shape.Type=1），获取所有图片占据的最大行号，
        返回截图起始行 = max(data_min_row, 最大图片底部行+1)
        """
        max_image_row = 0
        try:
            for shape in ws_com.Shapes:
                if shape.Type == 1:  # 1 表示图片 (xlPicture)
                    bottom_row = shape.BottomRightCell.Row
                    if bottom_row > max_image_row:
                        max_image_row = bottom_row
                        if self.verbose:
                            print(f"检测到图片，右下角行号: {bottom_row}")
        except Exception as e:
            if self.verbose:
                print(f"遍历图片时出错: {e}")
        start_row = max(data_min_row, max_image_row + 1)
        if self.verbose and max_image_row > 0:
            print(f"跳过顶部图片，起始行从 {data_min_row} 调整为 {start_row}")
        return start_row

    def _get_full_data_bounds(self, worksheet, ws_com=None):
        """
        获取数据边界，包括所有非空单元格和合并单元格。
        如果提供了 ws_com（COM 对象），则还会检测并跳过顶部图片。
        """
        # 先获取非空单元格的边界
        min_row = None
        max_row = None
        min_col = None
        max_col = None
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip():
                    if min_row is None or cell.row < min_row:
                        min_row = cell.row
                    if max_row is None or cell.row > max_row:
                        max_row = cell.row
                    if min_col is None or cell.column < min_col:
                        min_col = cell.column
                    if max_col is None or cell.column > max_col:
                        max_col = cell.column

        if min_row is None:
            min_row, max_row, min_col, max_col = 1, worksheet.max_row, 1, worksheet.max_column

        # 合并单元格扩展
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row < min_row:
                min_row = merged_range.min_row
            if merged_range.max_row > max_row:
                max_row = merged_range.max_row
            if merged_range.min_col < min_col:
                min_col = merged_range.min_col
            if merged_range.max_col > max_col:
                max_col = merged_range.max_col

        # 如果传入了 ws_com，则检测并跳过顶部图片
        if ws_com is not None:
            min_row = self._get_min_row_skip_images(ws_com, min_row)

        return min_row, max_row, min_col, max_col

    def _inject_and_run_vba(self, wb_com):
        try:
            vba_project = wb_com.VBProject
            module = None
            for comp in vba_project.VBComponents:
                if comp.Name == "AutoFitModule":
                    module = comp
                    break
            if module is None:
                module = vba_project.VBComponents.Add(1)
                module.Name = "AutoFitModule"
            code_module = module.CodeModule
            if code_module.CountOfLines == 0:
                code_module.AddFromString(self.vba_macro_code)
            wb_com.Application.Run("AutoFitAllMergedCells")
            if self.verbose:
                print("VBA宏执行成功")
            return True
        except Exception as e:
            if self.verbose:
                print(f"VBA宏注入或执行失败: {e}")
            return False

    def convert_full_table(self, excel_path: str, output_dir: str = "images", suffix: str = "") -> str:
        if not XLWINGS_AVAILABLE:
            raise RuntimeError("xlwings 不可用")
        abs_path = os.path.abspath(excel_path)
        if not os.path.exists(abs_path):
            raise FileNotFoundError(f"文件不存在: {abs_path}")

        # 用 openpyxl 获取基础边界（不含图片跳过）
        wb = openpyxl.load_workbook(abs_path, data_only=True)
        ws = wb.active
        min_row_temp, max_row, min_col, max_col = self._get_full_data_bounds(ws, ws_com=None)
        wb.close()

        # 打开工作簿，获取 COM 对象，用于检测图片
        wb_com = self.app.api.Workbooks.Open(abs_path, UpdateLinks=0)
        try:
            ws_com = wb_com.Worksheets(1)
            # 重新计算边界，此时会跳过图片
            min_row, max_row, min_col, max_col = self._get_full_data_bounds(ws, ws_com)
        except Exception as e:
            # 如果获取 COM 对象失败，回退到之前的边界
            min_row, max_row, min_col, max_col = min_row_temp, max_row, min_col, max_col
            if self.verbose:
                print(f"跳过图片检测失败: {e}")

        # 构造范围字符串
        range_str = f"{self._col_num_to_letter(min_col)}{min_row}:{self._col_num_to_letter(max_col)}{max_row}"
        if self.verbose:
            print(f"截图范围: {range_str}")
        os.makedirs(output_dir, exist_ok=True)
        base_name = os.path.splitext(os.path.basename(abs_path))[0]
        if suffix:
            output_path = os.path.join(output_dir, f"{base_name}_{suffix}.png")
        else:
            output_path = os.path.join(output_dir, f"{base_name}_full_table.png")

        try:
            data_range = ws_com.Range(range_str)

            if self.apply_formatting:
                # 列宽自适应
                data_range.Columns.AutoFit()

                # 单元格格式：水平居中 + 垂直居中
                data_range.HorizontalAlignment = -4108  # xlCenter
                data_range.VerticalAlignment = -4108  # xlCenter

                # ========== 字体增强 ==========
                #data_range.Font.Name = "微软雅黑"   # 或 "Arial"
                #data_range.Font.Bold = True
                # =============================

                # 增强边框：黑色、中等粗细、实线
                data_range.BorderAround(1, 2)  # 外边框
                data_range.Borders(11).LineStyle = 1  # 内部垂直
                data_range.Borders(11).Weight = 2
                data_range.Borders(12).LineStyle = 1  # 内部水平
                data_range.Borders(12).Weight = 2

                # 调整合并单元格行高
                self._inject_and_run_vba(wb_com)

            # 清空剪贴板
            try:
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.CloseClipboard()
            except:
                pass

            # 复制为图片
            data_range.CopyPicture(Appearance=1, Format=2)  # xlScreen, xlBitmap
            time.sleep(1.0)

            img = None
            for attempt in range(3):
                img = self._get_clipboard_image()
                if img:
                    break
                time.sleep(1.0)

            if img is None:
                raise RuntimeError("无法从剪贴板获取有效图片")

            img.save(output_path, 'PNG', quality=95)
            if self.verbose:
                print(f"🖼️ 已保存: {output_path}")
            return output_path
        finally:
            wb_com.Close(SaveChanges=False)

    def convert_full_table_with_options(self, excel_path: str, output_dir: str = "images") -> tuple:
        original_formatting = self.apply_formatting
        self.apply_formatting = True
        formatted_path = self.convert_full_table(excel_path, output_dir, suffix="formatted")
        self.apply_formatting = False
        raw_path = self.convert_full_table(excel_path, output_dir, suffix="raw")
        self.apply_formatting = original_formatting
        return formatted_path, raw_path