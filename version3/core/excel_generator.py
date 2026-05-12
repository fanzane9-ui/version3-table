#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel生成器 - 将JSON转换为结构化Excel（支持空行头数据归入上级）
"""

import json
import os
import re
from typing import List, Dict
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


# ========== 工具函数 ==========

def format_xlwings_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        else:
            return value.strftime("%Y-%m-%d")
    text = str(value)
    text = re.sub(r'\s+', ' ', text.replace('\u00A0', ' '))
    return text.strip()


def parse_sheet_with_xlwings(excel_path: str) -> List[Dict]:
    import xlwings as xw
    MAX_VOID = 2
    app = xw.App(visible=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        abs_path = os.path.abspath(excel_path)
        wb = app.books.open(abs_path)
        ws = wb.sheets[0]
        used_range = ws.used_range
        max_row = used_range.last_cell.row
        max_col = used_range.last_cell.column
        all_values = ws.range((1, 1), (max_row, max_col)).value
        if max_row == 1:
            all_values = [all_values]
        elif max_col == 1:
            all_values = [[v] for v in all_values]
        row_limit = 0
        for r in range(max_row):
            if any(v is not None and str(v).strip() != '' for v in all_values[r]):
                row_limit = r + 1
        row_limit = min(row_limit + MAX_VOID, max_row)
        col_limit = 0
        for c in range(max_col - 1, -1, -1):
            if any(all_values[r][c] is not None and str(all_values[r][c]).strip() != ''
                   for r in range(row_limit)):
                col_limit = c + 1
                break
        col_limit = min(col_limit + MAX_VOID, max_col)
        wb.close()
        wb_openpyxl = load_workbook(excel_path, data_only=False)
        ws_openpyxl = wb_openpyxl.active
        merged_dict = {}
        result = []
        for mrg in ws_openpyxl.merged_cells.ranges:
            min_col, min_row, max_col_mrg, max_row_mrg = mrg.bounds
            if min_row > row_limit or min_col > col_limit:
                continue
            rows = list(range(min_row - 1, max_row_mrg))
            cols = list(range(min_col - 1, max_col_mrg))
            cell_value = all_values[min_row - 1][min_col - 1]
            value = format_xlwings_value(cell_value)
            result.append({"words": value, "rows": rows, "columns": cols})
            for r in rows:
                for c in cols:
                    merged_dict[(r, c)] = True
        for r in range(row_limit):
            for c in range(col_limit):
                if (r, c) in merged_dict:
                    continue
                cell_value = all_values[r][c]
                value = format_xlwings_value(cell_value)
                result.append({"words": value, "rows": [r], "columns": [c]})
        result.sort(key=lambda item: (min(item["rows"]), min(item["columns"])))
        return result
    finally:
        app.quit()


def convert_excel_to_merged_json(excel_path: str, output_path: str = None, use_xlwings: bool = True) -> str:
    if output_path is None:
        base_name = os.path.splitext(os.path.basename(excel_path))[0]
        output_dir = os.path.dirname(excel_path) if os.path.dirname(excel_path) else '.'
        output_path = os.path.join(output_dir, f"{base_name}_merged.json")
    if use_xlwings:
        try:
            import xlwings as xw
            data = parse_sheet_with_xlwings(excel_path)
        except ImportError:
            print("⚠️ xlwings未安装，回退到openpyxl")
            data = []
        except Exception as e:
            print(f"⚠️ xlwings读取失败: {e}")
            data = []
    else:
        print("⚠️ 使用openpyxl模式")
        data = []
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[OK] merged.json已生成: {output_path} (共{len(data)}条记录)")
    return output_path


# ========== JSON 转 Excel 核心 ==========

class TreeNode:
    def __init__(self, words, rows, columns, type):
        self.words = words
        self.rows = rows
        self.columns = columns
        self.type = type


def safe_json_load(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return json.load(file)
    except UnicodeDecodeError:
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as file:
                return json.load(file)
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='gbk') as file:
                return json.load(file)


def build_column_paths(header_nodes):
    """
    基于几何包含关系构建列头路径。
    参数：header_nodes - list of TreeNode (type='column_header')
    返回：(paths, leaf_nodes)  paths为每个叶子节点的完整路径（用|分隔），leaf_nodes为叶子节点列表
    """
    if not header_nodes:
        return [], []
    # 按行号、列号排序
    header_nodes.sort(key=lambda x: (x.rows[0], x.columns[0]))

    # 标记叶子：没有其他节点在其下方且列范围被其包含
    is_leaf = [True] * len(header_nodes)
    for i, node in enumerate(header_nodes):
        for j, other in enumerate(header_nodes):
            if i == j:
                continue
            # 如果 other 在 node 下方且列范围被 node 包含，则 node 不是叶子
            if (other.rows[0] > node.rows[0] and
                    node.columns[0] <= other.columns[0] and
                    node.columns[-1] >= other.columns[-1]):
                is_leaf[i] = False
                break
    leaf_nodes = [header_nodes[i] for i, leaf in enumerate(is_leaf) if leaf]

    # 为每个叶子节点找祖先
    paths = []
    for leaf in leaf_nodes:
        ancestors = []
        for other in header_nodes:
            if other is leaf:
                continue
            # 祖先必须在叶子节点上方，且列范围包含叶子的列范围
            if (other.rows[0] < leaf.rows[0] and
                    other.columns[0] <= leaf.columns[0] and
                    other.columns[-1] >= leaf.columns[-1]):
                ancestors.append(other)
        # 按行号排序（从上到下）
        ancestors.sort(key=lambda x: x.rows[0])
        path_words = [anc.words for anc in ancestors] + [leaf.words]
        paths.append('|'.join(path_words))
    return paths, leaf_nodes


def build_row_paths(header_nodes):
    """
    构建行头路径，忽略空单元格（words为空或仅空白），但保留叶子节点的位置占位。
    空行头不会出现在路径中，其下方的数据会归入上一个非空行头。
    """
    if not header_nodes:
        return [], []
    # 过滤掉空单元格（words为空或仅空白）
    non_empty_nodes = [node for node in header_nodes if node.words and node.words.strip()]
    # 按列号、行号排序
    non_empty_nodes.sort(key=lambda x: (x.columns[0], x.rows[0]))

    if not non_empty_nodes:
        return [], []

    # 构建父子关系（仅基于非空节点）
    children = {node: [] for node in non_empty_nodes}
    parents = {node: [] for node in non_empty_nodes}
    for i, node in enumerate(non_empty_nodes):
        for j, other in enumerate(non_empty_nodes):
            if i == j:
                continue
            # other 在 node 右侧且行范围被 node 包含
            if (other.columns[0] > node.columns[0] and
                    node.rows[0] <= other.rows[0] and
                    node.rows[-1] >= other.rows[-1]):
                children[node].append(other)
                parents[other].append(node)

    leaf_nodes = [node for node in non_empty_nodes if not children[node]]

    def get_path(node):
        path = [node.words]
        p = node
        while parents[p]:
            p = parents[p][0]
            path.insert(0, p.words)
        return '|'.join(path)

    paths = [get_path(leaf) for leaf in leaf_nodes]
    return paths, leaf_nodes


def convert_json_to_structured_excel(json_path: str, output_path: str = None, use_qwen: bool = False) -> str:
    if output_path is None:
        base_name = os.path.splitext(os.path.basename(json_path))[0]
        output_path = f"{base_name}_structured.xlsx"

    words_block_list = safe_json_load(json_path)
    print(f"[OK] JSON文件加载成功: {json_path}")

    # 创建节点列表并分类
    col_header_nodes = []
    row_header_nodes = []
    data_nodes = []
    for block in words_block_list:
        node = TreeNode(block['words'], block['rows'], block['columns'], block['type'])
        if block['type'] == 'column_header':
            col_header_nodes.append(node)
        elif block['type'] == 'row_header':
            row_header_nodes.append(node)
        else:  # data
            data_nodes.append(node)

    # 构建列头路径
    col_paths, col_leaves = build_column_paths(col_header_nodes)
    # 构建行头路径
    row_paths, row_leaves = build_row_paths(row_header_nodes)

    # 数据节点排序
    sorted_data = sorted(data_nodes, key=lambda x: (x.rows[0], x.columns[0]))

    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "结构化表格"

    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='left', vertical='center')

    has_row_headers = len(row_paths) > 0
    has_col_headers = len(col_paths) > 0

    col_start = 2 if has_row_headers else 1
    row_start = 2 if has_col_headers else 1

    # 写入列标题
    if has_col_headers:
        for idx, path in enumerate(col_paths):
            cell = ws.cell(row=1, column=col_start + idx)
            cell.value = path.replace('|', ' > ')
            cell.font = header_font
            cell.alignment = header_alignment

    # 写入行标题
    if has_row_headers:
        for idx, path in enumerate(row_paths):
            row_num = row_start + idx
            cell = ws.cell(row=row_num, column=1)
            cell.value = path.replace('|', ' > ')
            cell.font = header_font
            cell.alignment = header_alignment
        col_start = 2  # 行标题固定占第一列

    # 列映射：叶子节点覆盖的原始列号 -> Excel列号
    col_mapping = {}
    if has_col_headers:
        for idx, leaf in enumerate(col_leaves):
            excel_col = col_start + idx
            for orig_col in leaf.columns:
                col_mapping[orig_col] = excel_col
    else:
        # 没有列头时，直接映射数据列（排除行头占用的列）
        row_header_cols = set()
        if has_row_headers:
            for leaf in row_leaves:
                row_header_cols.update(leaf.columns)
        all_cols = sorted(set(col for node in data_nodes for col in node.columns))
        data_cols = [col for col in all_cols if col not in row_header_cols]
        for idx, orig_col in enumerate(data_cols):
            col_mapping[orig_col] = col_start + idx

    # 行映射：叶子节点覆盖的原始行号 -> Excel行号
    row_mapping = {}
    if has_row_headers:
        # 建立叶子节点覆盖的行号集合
        leaf_row_sets = [set(leaf.rows) for leaf in row_leaves]
        # 收集所有数据行
        all_data_rows = sorted(set(row for node in data_nodes for row in node.rows))
        # 为每个数据行寻找对应的叶子节点
        for orig_row in all_data_rows:
            matched = False
            for idx, leaf in enumerate(row_leaves):
                if orig_row in leaf.rows:
                    row_mapping[orig_row] = row_start + idx
                    matched = True
                    break
            if not matched:
                # 未匹配：归入最近的上一个叶子节点（按行号）
                # 找到小于当前行的最大叶子节点行
                best_idx = -1
                best_row = -1
                for idx, leaf in enumerate(row_leaves):
                    if leaf.rows[0] <= orig_row and leaf.rows[0] > best_row:
                        best_row = leaf.rows[0]
                        best_idx = idx
                if best_idx != -1:
                    row_mapping[orig_row] = row_start + best_idx
                else:
                    # 实在没有，放到第一个
                    row_mapping[orig_row] = row_start
    else:
        # 没有行头时，找出所有数据行（排除表头行）
        max_header_row = -1
        for node in col_header_nodes:
            max_header_row = max(max_header_row, max(node.rows))
        all_data_rows = sorted(set(row for node in data_nodes for row in node.rows if row > max_header_row))
        for idx, orig_row in enumerate(all_data_rows):
            row_mapping[orig_row] = row_start + idx

    # 填充数据
    from openpyxl.cell.cell import MergedCell
    for node in sorted_data:
        orig_rows = node.rows
        orig_cols = node.columns
        excel_rows = [row_mapping.get(r) for r in orig_rows if row_mapping.get(r)]
        excel_cols = [col_mapping.get(c) for c in orig_cols if col_mapping.get(c)]
        if excel_rows and excel_cols:
            start_row = min(excel_rows)
            end_row = max(excel_rows)
            start_col = min(excel_cols)
            end_col = max(excel_cols)
            cell = ws.cell(row=start_row, column=start_col)
            if isinstance(cell, MergedCell):
                continue
            if node.words:
                try:
                    # 尝试转换为数字
                    if node.words.replace('.', '').replace('-', '').replace(' ', '').isdigit():
                        cell.value = float(node.words) if '.' in node.words else int(node.words)
                        cell.number_format = '0.##########'
                    else:
                        cell.value = node.words
                except:
                    cell.value = node.words
            else:
                cell.value = ""
            cell.alignment = data_alignment
            if start_row != end_row or start_col != end_col:
                ws.merge_cells(start_row=start_row, start_column=start_col,
                               end_row=end_row, end_column=end_col)

    # 调整列宽
    if has_row_headers:
        ws.column_dimensions['A'].width = 25
    for col_idx in range(col_start, col_start + max(1, len(col_paths) if has_col_headers else len(col_mapping))):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    wb.save(output_path)
    print(f"[OK] 结构化表格已保存: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    import glob

    print("=" * 60)
    print("Excel Generator 批量测试")
    print("=" * 60)
    final_dir = "output/final/json"
    if not os.path.exists(final_dir):
        print(f"[ERROR] 目录不存在: {final_dir}")
        sys.exit(1)
    test_files = glob.glob(os.path.join(final_dir, "*_final.json"))
    if not test_files:
        print(f"[ERROR] 在 {final_dir} 目录下未找到 *_final.json 文件")
        sys.exit(1)
    print(f"\n找到 {len(test_files)} 个测试文件:")
    for i, f in enumerate(test_files, 1):
        print(f"  {i}. {os.path.basename(f)}")
    print()
    use_qwen = len(sys.argv) > 1 and sys.argv[1] == "--use-qwen"
    if use_qwen:
        print("🤖 使用 Qwen 智能判断左上角单元格")
    else:
        print("📋 使用默认布局 (不使用 Qwen)")
    print("=" * 60)
    print()
    success_count = 0
    failed_count = 0
    results = []
    for idx, test_file in enumerate(test_files, 1):
        file_name = os.path.basename(test_file)
        print(f"[{idx}/{len(test_files)}] 处理: {file_name}")
        print("-" * 60)
        base_name = os.path.basename(test_file)
        excel_dir = "output/final/excel"
        os.makedirs(excel_dir, exist_ok=True)
        if use_qwen:
            output_file = os.path.join(excel_dir, base_name.replace('_final.json', '_structured_qwen.xlsx'))
        else:
            output_file = os.path.join(excel_dir, base_name.replace('_final.json', '_structured.xlsx'))
        try:
            result = convert_json_to_structured_excel(test_file, output_file, use_qwen=use_qwen)
            print(f"✅ 成功: {os.path.basename(result)}")
            success_count += 1
            results.append({"file": file_name, "status": "成功", "output": os.path.basename(result)})
        except Exception as e:
            print(f"❌ 失败: {e}")
            failed_count += 1
            results.append({"file": file_name, "status": "失败", "error": str(e)})
        print()
    print("=" * 60)
    print("测试汇总")
    print("=" * 60)
    print(f"总计: {len(test_files)} 个文件")
    print(f"成功: {success_count} 个")
    print(f"失败: {failed_count} 个")
    print()
    if failed_count > 0:
        print("失败文件列表:")
        for r in results:
            if r["status"] == "失败":
                print(f"  ❌ {r['file']}: {r['error']}")
        print()
    if success_count > 0:
        print("成功文件列表:")
        for r in results:
            if r["status"] == "成功":
                print(f"  ✅ {r['file']} → {r['output']}")
    print("=" * 60)
    print("测试完成！")
    print("=" * 60)