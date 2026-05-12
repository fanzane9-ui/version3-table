#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 提取器
从人工矫正后的 Excel 文件中提取单元格信息，包括文本、位置和类型
输出 JSON 格式，每个单元格一个对象
"""

import os
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelExtractor:
    """从标注后的 Excel 提取单元格信息"""

    COLOR_TYPE_MAP = {
        'FF0000': 'column_header',
        '00B050': 'row_header',
    }

    def __init__(self, debug=False):
        self.debug = debug

    def _get_cell_type(self, cell):
        """根据单元格填充颜色判断类型"""
        if self.debug:
            print(f"【调试】单元格 ({cell.row},{cell.column}) 值='{cell.value}'")
            print(f"        fill 类型: {type(cell.fill)}")

        fg_color = None
        if cell.fill:
            try:
                fg_color = cell.fill.fgColor
            except AttributeError:
                pass

        if fg_color:
            if self.debug:
                print(f"        fg_color 对象: {fg_color}")
            rgb = fg_color.rgb
            if self.debug:
                print(f"        rgb 原始值: {rgb} (类型: {type(rgb)})")

            if rgb is not None:
                if hasattr(rgb, 'rgb'):
                    rgb_str = rgb.rgb
                else:
                    rgb_str = str(rgb)

                if rgb_str and len(rgb_str) >= 6:
                    color = rgb_str[-6:].upper()
                    if self.debug:
                        print(f"        提取的颜色: '{color}'")
                    if color in self.COLOR_TYPE_MAP:
                        return self.COLOR_TYPE_MAP[color]
                elif self.debug:
                    print(f"        rgb 字符串无效: '{rgb_str}'")
            elif self.debug:
                print(f"        rgb 为 None")
        elif self.debug:
            print(f"        fg_color 为 None")

        return 'data'

    def _get_merged_ranges(self, worksheet):
        """获取合并单元格区域映射"""
        merged_map = {}
        for merged_range in worksheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            merged_map[(min_row, min_col)] = (min_row, max_row, min_col, max_col)
        return merged_map

    def extract_from_workbook(self, workbook_path):
        """从 Excel 文件中提取所有单元格信息"""
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb.active
        merged_map = self._get_merged_ranges(ws)
        cells_info = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or str(cell.value).strip() == '':
                    continue

                # 跳过非左上角的合并单元格
                is_non_upper = False
                for (tl_row, tl_col), (min_r, max_r, min_c, max_c) in merged_map.items():
                    if min_r <= cell.row <= max_r and min_c <= cell.column <= max_c:
                        if cell.row != tl_row or cell.column != tl_col:
                            is_non_upper = True
                            break
                if is_non_upper:
                    continue

                if (cell.row, cell.column) in merged_map:
                    min_row, max_row, min_col, max_col = merged_map[(cell.row, cell.column)]
                    rows = list(range(min_row - 1, max_row))
                    cols = list(range(min_col - 1, max_col))
                else:
                    rows = [cell.row - 1]
                    cols = [cell.column - 1]

                cell_text = str(cell.value).strip()
                cell_type = self._get_cell_type(cell)

                cells_info.append({
                    "words": cell_text,
                    "rows": rows,
                    "columns": cols,
                    "type": cell_type
                })

        wb.close()
        return cells_info

    def save_to_json(self, cells_info, output_path):
        """将单元格信息保存为 JSON 文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(cells_info, f, ensure_ascii=False, indent=2)

    def process_file(self, excel_path, output_json_path):
        """处理单个文件：提取并保存 JSON"""
        print(f"📄 处理: {os.path.basename(excel_path)}")
        cells_info = self.extract_from_workbook(excel_path)
        self.save_to_json(cells_info, output_json_path)
        print(f"   ✅ 已提取 {len(cells_info)} 个单元格 -> {output_json_path}")
        return cells_info


def batch_extract(input_folder, output_folder, debug=False):
    """
    批量处理文件夹中的所有 Excel 文件
    input_folder: 包含标注后 Excel 的文件夹（如 output/table_mark）
    output_folder: 输出 JSON 的文件夹（如 output/final）
    debug: 是否输出调试信息
    """
    os.makedirs(output_folder, exist_ok=True)
    extractor = ExcelExtractor(debug=debug)

    excel_files = [f for f in os.listdir(input_folder) if f.endswith('.xlsx')]
    if not excel_files:
        print("❌ 未找到 Excel 文件")
        return

    for filename in excel_files:
        base = os.path.splitext(filename)[0]
        excel_path = os.path.join(input_folder, filename)
        json_path = os.path.join(output_folder, f"{base}.json")
        extractor.process_file(excel_path, json_path)